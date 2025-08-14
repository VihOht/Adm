import csv
import json
from datetime import datetime
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_POST

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from finance_manager.models import ExpenseCategory, Expenses, IncomeCategorys, Incomes

# Create your views here.


@login_required
def dashboard(request):
    """Finance manager dashboard view"""
    expenses = Expenses.objects.filter(user=request.user).order_by("-spent_at")[:5]
    expenses_this_month = Expenses.objects.filter(
        user=request.user,
        spent_at__year=datetime.now().year,
        spent_at__month=datetime.now().month,
    )
    incomes = Incomes.objects.filter(user=request.user).order_by("-received_at")[:5]
    expense_categories = ExpenseCategory.objects.filter(user=request.user)
    income_categories = IncomeCategorys.objects.filter(user=request.user)

    total_expenses = sum([x.amount for x in Expenses.objects.filter(user=request.user)])
    total_incomes = sum([x.amount for x in Incomes.objects.filter(user=request.user)])

    range_of_days = "Sem gastos neste mês"
    total_expenses_this_month = 0
    daily_mean = 0
    if expenses_this_month:
        range_of_days = f"Últimos {datetime.now().day} dias"
        total_expenses_this_month = sum([x.amount for x in expenses_this_month])
        daily_mean = total_expenses_this_month // datetime.now().day

    total_amount = total_incomes - total_expenses

    return render(
        request,
        "finance_manager/dashboard.html",
        {
            "expenses": expenses,
            "incomes": incomes,
            "expense_categories": expense_categories,
            "income_categories": income_categories,
            "categories": expense_categories,  # For backward compatibility
            "total_expenses_this_month": f"{total_expenses_this_month // 100},{total_expenses_this_month % 100:02n}",
            "total_amount": f"{total_amount // 100},{total_amount % 100:02n}",
            "range_of_days": range_of_days,
            "daily_mean": f"{daily_mean // 100},{daily_mean % 100:02n}",
        },
    )


@login_required
def transactions(request):
    """Transactions list view showing all incomes and expenses"""
    # Get all user's transactions
    expenses = Expenses.objects.filter(user=request.user).order_by("-spent_at")
    incomes = Incomes.objects.filter(user=request.user).order_by("-received_at")

    # Create combined transactions list ordered by date (most recent first)
    all_transactions = []

    # Add expenses with type indicator
    for expense in expenses:
        all_transactions.append(
            {
                "type": "expense",
                "id": expense.id,
                "description": expense.description,
                "detailed_description": expense.detailed_description,
                "amount": expense.amount,
                "date": expense.spent_at,
                "category": expense.category,
                "object": expense,
            }
        )

    # Add incomes with type indicator
    for income in incomes:
        all_transactions.append(
            {
                "type": "income",
                "id": income.id,
                "description": income.description,
                "detailed_description": income.detailed_description,
                "amount": income.amount,
                "date": income.received_at,
                "category": income.category,
                "object": income,
            }
        )

    # Sort all transactions by date (most recent first)
    all_transactions.sort(key=lambda x: x["date"], reverse=True)

    # Get categories for editing
    expense_categories = ExpenseCategory.objects.filter(user=request.user)
    income_categories = IncomeCategorys.objects.filter(user=request.user)

    # Calculate totals
    total_expenses = sum([x.amount for x in expenses])
    total_incomes = sum([x.amount for x in incomes])
    balance = total_incomes - total_expenses

    # Format amounts for display
    def format_amount(amount):
        return f"{amount // 100},{amount % 100:02n}"

    return render(
        request,
        "finance_manager/transactions.html",
        {
            "expenses": expenses,
            "incomes": incomes,
            "all_transactions": all_transactions,
            "expense_categories": expense_categories,
            "income_categories": income_categories,
            "total_expenses": format_amount(total_expenses),
            "total_incomes": format_amount(total_incomes),
            "balance": format_amount(balance),
            "balance_raw": balance,
        },
    )


@login_required
@require_POST
def create_expense(request):
    """Create a new expense via AJAX"""
    try:
        data = json.loads(request.body)

        # Get or create category
        category = None
        if data.get("category_id"):
            try:
                category = ExpenseCategory.objects.get(
                    id=data["category_id"], user=request.user
                )
            except ExpenseCategory.DoesNotExist:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Categoria não encontrada ou não pertence ao usuário.",
                    }
                )

        # Convert amount from cents to integer (assuming input is in reais)
        amount_in_cents = int(float(data["amount"]) * 100)

        # Parse date
        spent_date = datetime.strptime(data["spent_at"], "%Y-%m-%d").date()

        # Create expense
        expense = Expenses.objects.create(
            user=request.user,
            category=category,
            description=data["description"],
            detailed_description=data.get("detailed_description", ""),
            amount=amount_in_cents,
            spent_at=spent_date,
        )

        return JsonResponse(
            {
                "success": True,
                "message": "Gasto adicionado com sucesso!",
                "expense_id": expense.id,
            }
        )

    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao criar gasto: {str(e)}"}
        )


@login_required
@require_POST
def create_category(request):
    """Create a new expense category via AJAX"""
    try:
        data = json.loads(request.body)

        category = ExpenseCategory.objects.create(
            user=request.user,
            name=data["name"],
            description=data.get("description", ""),
            color=data.get("color", "#FFFFFF"),
        )

        return JsonResponse(
            {
                "success": True,
                "message": "Categoria criada com sucesso!",
                "category": {
                    "id": category.id,
                    "name": category.name,
                    "description": category.description,
                    "color": category.color,
                },
            }
        )

    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao criar categoria: {str(e)}"}
        )


@login_required
@require_POST
def edit_category(request, category_id):
    """Edit an existing expense category via AJAX"""
    try:
        category = get_object_or_404(ExpenseCategory, id=category_id, user=request.user)
        data = json.loads(request.body)

        # Update category
        category.name = data["name"]
        category.description = data.get("description", "")
        category.color = data.get("color", "#FFFFFF")
        category.save()

        return JsonResponse(
            {
                "success": True,
                "message": "Categoria atualizada com sucesso!",
                "category": {
                    "id": category.id,
                    "name": category.name,
                    "description": category.description,
                    "color": category.color,
                },
            }
        )

    except ExpenseCategory.DoesNotExist:
        return JsonResponse({"success": False, "message": "Categoria não encontrada"})
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao atualizar categoria: {str(e)}"}
        )


@login_required
@require_POST
def delete_category(request, category_id):
    """Delete an existing expense category and all related expenses via AJAX"""
    try:
        category = get_object_or_404(ExpenseCategory, id=category_id, user=request.user)

        # Count related expenses before deletion
        related_expenses_count = Expenses.objects.filter(category=category).count()

        category_name = category.name

        # Delete the category (this will cascade delete related expenses due to ForeignKey)
        category.delete()

        message = f"Categoria '{category_name}' excluída com sucesso!"
        if related_expenses_count > 0:
            message += f" {related_expenses_count} gasto(s) relacionado(s) também foram excluídos."

        return JsonResponse(
            {
                "success": True,
                "message": message,
                "deleted_expenses_count": related_expenses_count,
            }
        )

    except ExpenseCategory.DoesNotExist:
        return JsonResponse({"success": False, "message": "Categoria não encontrada"})
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao excluir categoria: {str(e)}"}
        )


@login_required
@require_POST
def edit_income_category(request, category_id):
    """Edit an existing income category via AJAX"""
    try:
        category = get_object_or_404(IncomeCategorys, id=category_id, user=request.user)
        data = json.loads(request.body)

        # Update category
        category.name = data["name"]
        category.description = data.get("description", "")
        category.color = data.get("color", "#FFFFFF")
        category.save()

        return JsonResponse(
            {
                "success": True,
                "message": "Categoria de renda atualizada com sucesso!",
                "category": {
                    "id": category.id,
                    "name": category.name,
                    "description": category.description,
                    "color": category.color,
                },
            }
        )

    except IncomeCategorys.DoesNotExist:
        return JsonResponse(
            {"success": False, "message": "Categoria de renda não encontrada"}
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "message": f"Erro ao atualizar categoria de renda: {str(e)}",
            }
        )


@login_required
@require_POST
def delete_income_category(request, category_id):
    """Delete an existing income category and all related incomes via AJAX"""
    try:
        category = get_object_or_404(IncomeCategorys, id=category_id, user=request.user)

        # Count related incomes before deletion
        related_incomes_count = Incomes.objects.filter(category=category).count()

        category_name = category.name

        # Delete the category (this will cascade delete related incomes due to ForeignKey)
        category.delete()

        message = f"Categoria de renda '{category_name}' excluída com sucesso!"
        if related_incomes_count > 0:
            message += f" {related_incomes_count} receita(s) relacionada(s) também foram excluídas."

        return JsonResponse(
            {
                "success": True,
                "message": message,
                "deleted_incomes_count": related_incomes_count,
            }
        )

    except IncomeCategorys.DoesNotExist:
        return JsonResponse(
            {"success": False, "message": "Categoria de renda não encontrada"}
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "message": f"Erro ao excluir categoria de renda: {str(e)}",
            }
        )


@login_required
@require_POST
def edit_expense(request, expense_id):
    """Edit an existing expense via AJAX"""
    try:
        expense = get_object_or_404(Expenses, id=expense_id, user=request.user)
        data = json.loads(request.body)

        # Get or update category
        category = None
        if data.get("category_id"):
            try:
                category = ExpenseCategory.objects.get(
                    id=data["category_id"], user=request.user
                )
            except ExpenseCategory.DoesNotExist:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Categoria não encontrada ou não pertence ao usuário.",
                    }
                )

        # Convert amount from cents to integer (assuming input is in reais)
        amount_in_cents = int(float(data["amount"]) * 100)

        # Parse date
        spent_date = datetime.strptime(data["spent_at"], "%Y-%m-%d").date()

        # Update expense
        expense.category = category
        expense.description = data["description"]
        expense.detailed_description = data.get("detailed_description", "")
        expense.amount = amount_in_cents
        expense.spent_at = spent_date
        expense.save()

        return JsonResponse(
            {
                "success": True,
                "message": "Gasto atualizado com sucesso!",
                "expense_id": expense.id,
            }
        )

    except Expenses.DoesNotExist:
        return JsonResponse({"success": False, "message": "Gasto não encontrado"})
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao atualizar gasto: {str(e)}"}
        )


@login_required
@require_POST
def delete_expense(request, expense_id):
    """Delete an existing expense via AJAX"""
    try:
        expense = get_object_or_404(Expenses, id=expense_id, user=request.user)
        expense.delete()

        return JsonResponse(
            {
                "success": True,
                "message": "Gasto excluído com sucesso!",
            }
        )

    except Expenses.DoesNotExist:
        return JsonResponse({"success": False, "message": "Gasto não encontrado"})
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao excluir gasto: {str(e)}"}
        )


@login_required
@require_POST
def create_income(request):
    """Create a new income via AJAX"""
    try:
        data = json.loads(request.body)

        # Get or create category
        category = None
        if data.get("category_id"):
            try:
                category = IncomeCategorys.objects.get(
                    id=data["category_id"], user=request.user
                )
            except IncomeCategorys.DoesNotExist:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Categoria não encontrada ou não pertence ao usuário.",
                    }
                )

        # Convert amount from cents to integer (assuming input is in reais)
        amount_in_cents = int(float(data["amount"]) * 100)

        # Parse date
        received_date = datetime.strptime(data["received_at"], "%Y-%m-%d").date()

        # Create income
        income = Incomes.objects.create(
            user=request.user,
            category=category,
            description=data["description"],
            detailed_description=data.get("detailed_description", ""),
            amount=amount_in_cents,
            received_at=received_date,
        )

        return JsonResponse(
            {
                "success": True,
                "message": "Receita adicionada com sucesso!",
                "income_id": income.id,
            }
        )

    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao criar receita: {str(e)}"}
        )


@login_required
@require_POST
def create_income_category(request):
    """Create a new income category via AJAX"""
    try:
        data = json.loads(request.body)

        category = IncomeCategorys.objects.create(
            user=request.user,
            name=data["name"],
            description=data.get("description", ""),
            color=data.get("color", "#FFFFFF"),
        )

        return JsonResponse(
            {
                "success": True,
                "message": "Categoria de receita criada com sucesso!",
                "category": {
                    "id": category.id,
                    "name": category.name,
                    "description": category.description,
                    "color": category.color,
                },
            }
        )

    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao criar categoria: {str(e)}"}
        )


@login_required
@require_POST
def edit_income(request, income_id):
    """Edit an existing income via AJAX"""
    try:
        income = get_object_or_404(Incomes, id=income_id, user=request.user)
        data = json.loads(request.body)

        # Get or update category
        category = None
        if data.get("category_id"):
            try:
                category = IncomeCategorys.objects.get(
                    id=data["category_id"], user=request.user
                )
            except IncomeCategorys.DoesNotExist:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Categoria não encontrada ou não pertence ao usuário.",
                    }
                )

        # Convert amount from cents to integer (assuming input is in reais)
        amount_in_cents = int(float(data["amount"]) * 100)

        # Parse date
        received_date = datetime.strptime(data["received_at"], "%Y-%m-%d").date()

        # Update income
        income.category = category
        income.description = data["description"]
        income.detailed_description = data.get("detailed_description", "")
        income.amount = amount_in_cents
        income.received_at = received_date
        income.save()

        return JsonResponse(
            {
                "success": True,
                "message": "Receita atualizada com sucesso!",
                "income_id": income.id,
            }
        )

    except Incomes.DoesNotExist:
        return JsonResponse({"success": False, "message": "Receita não encontrada"})
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao atualizar receita: {str(e)}"}
        )


@login_required
@require_POST
def delete_income(request, income_id):
    """Delete an existing income via AJAX"""
    try:
        income = get_object_or_404(Incomes, id=income_id, user=request.user)
        income.delete()

        return JsonResponse(
            {
                "success": True,
                "message": "Receita excluída com sucesso!",
            }
        )

    except Incomes.DoesNotExist:
        return JsonResponse({"success": False, "message": "Receita não encontrada"})
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao excluir receita: {str(e)}"}
        )


@login_required
def export_financial_data_json(request):
    """Export user's financial data as JSON file"""

    user = request.user

    # Prepare data structure
    data = {
        "exported_at": datetime.now().isoformat(),
        "username": user.username,
        "expense_categories": [],
        "income_categories": [],
        "expenses": [],
        "incomes": [],
    }

    # Export expense categories
    for category in ExpenseCategory.objects.filter(user=user):
        data["expense_categories"].append(
            {
                "name": category.name,
                "description": category.description,
                "color": category.color,
            }
        )

    # Export income categories
    for category in IncomeCategorys.objects.filter(user=user):
        data["income_categories"].append(
            {
                "name": category.name,
                "description": category.description,
                "color": category.color,
            }
        )

    # Export expenses
    for expense in Expenses.objects.filter(user=user).order_by("-spent_at"):
        data["expenses"].append(
            {
                "category": expense.category.name if expense.category else None,
                "spent_at": expense.spent_at.strftime("%Y-%m-%d"),
                "description": expense.description,
                "detailed_description": expense.detailed_description,
                "amount": expense.amount,
                "created_at": expense.created_at.isoformat(),
            }
        )

    # Export incomes
    for income in Incomes.objects.filter(user=user).order_by("-received_at"):
        data["incomes"].append(
            {
                "category": income.category.name if income.category else None,
                "received_at": income.received_at.strftime("%Y-%m-%d"),
                "description": income.description,
                "detailed_description": income.detailed_description,
                "amount": income.amount,
                "created_at": income.created_at.isoformat(),
            }
        )

    # Create HTTP response with JSON data
    response = HttpResponse(
        json.dumps(data, indent=2, ensure_ascii=False),
        content_type="application/json",
    )

    # Set download headers
    filename = f"fd_{user.username}_{datetime.now().strftime('%Y_%m_%d')}.json"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
def export_financial_data_csv(request):
    """Export user's financial data as CSV file"""
    user = request.user

    # Create HTTP response with CSV content
    response = HttpResponse(content_type="text/csv")
    filename = f"fd_{user.username}_{datetime.now().strftime('%Y_%m_%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Write header information
    writer.writerow(["Exported At", "Username"])
    writer.writerow([datetime.now().isoformat(), user.username])
    writer.writerow([])  # Empty row for separation

    # Write expense categories
    writer.writerow(["EXPENSE CATEGORIES"])
    writer.writerow(["Name", "Description", "Color"])
    for category in ExpenseCategory.objects.filter(user=user):
        writer.writerow([category.name, category.description, category.color])

    writer.writerow([])  # Empty row for separation

    # Write income categories
    writer.writerow(["INCOME CATEGORIES"])
    writer.writerow(["Name", "Description", "Color"])
    for category in IncomeCategorys.objects.filter(user=user):
        writer.writerow([category.name, category.description, category.color])

    writer.writerow([])  # Empty row for separation

    # Write expenses
    writer.writerow(["EXPENSES"])
    writer.writerow(
        [
            "Category",
            "Date",
            "Description",
            "Detailed Description",
            "Amount (cents)",
            "Created At",
        ]
    )
    for expense in Expenses.objects.filter(user=user).order_by("-spent_at"):
        writer.writerow(
            [
                expense.category.name if expense.category else "",
                expense.spent_at.strftime("%Y-%m-%d"),
                expense.description,
                expense.detailed_description,
                expense.amount,
                expense.created_at.isoformat(),
            ]
        )

    writer.writerow([])  # Empty row for separation

    # Write incomes
    writer.writerow(["INCOMES"])
    writer.writerow(
        [
            "Category",
            "Date",
            "Description",
            "Detailed Description",
            "Amount (cents)",
            "Created At",
        ]
    )
    for income in Incomes.objects.filter(user=user).order_by("-received_at"):
        writer.writerow(
            [
                income.category.name if income.category else "",
                income.received_at.strftime("%Y-%m-%d"),
                income.description,
                income.detailed_description,
                income.amount,
                income.created_at.isoformat(),
            ]
        )

    return response


@login_required
def export_financial_data_excel(request):
    """Export user's financial data as Excel file"""
    if not EXCEL_AVAILABLE:
        return JsonResponse(
            {
                "success": False,
                "message": "Excel export não disponível. Instale openpyxl: pip install openpyxl",
            }
        )

    user = request.user

    # Create workbook and worksheets
    wb = Workbook()

    # Remove default worksheet
    wb.remove(wb.active)

    # Create worksheets
    info_ws = wb.create_sheet("Info")
    exp_cat_ws = wb.create_sheet("Expense Categories")
    inc_cat_ws = wb.create_sheet("Income Categories")
    expenses_ws = wb.create_sheet("Expenses")
    incomes_ws = wb.create_sheet("Incomes")

    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )

    # Info worksheet
    info_ws["A1"] = "Export Information"
    info_ws["A1"].font = header_font
    info_ws["A1"].fill = header_fill
    info_ws["A2"] = "Exported At:"
    info_ws["B2"] = datetime.now().isoformat()
    info_ws["A3"] = "Username:"
    info_ws["B3"] = user.username

    # Expense Categories worksheet
    exp_cat_headers = ["Name", "Description", "Color"]
    for col, header in enumerate(exp_cat_headers, 1):
        cell = exp_cat_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, category in enumerate(ExpenseCategory.objects.filter(user=user), 2):
        exp_cat_ws.cell(row=row, column=1, value=category.name)
        exp_cat_ws.cell(row=row, column=2, value=category.description)
        exp_cat_ws.cell(row=row, column=3, value=category.color)

    # Income Categories worksheet
    inc_cat_headers = ["Name", "Description", "Color"]
    for col, header in enumerate(inc_cat_headers, 1):
        cell = inc_cat_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, category in enumerate(IncomeCategorys.objects.filter(user=user), 2):
        inc_cat_ws.cell(row=row, column=1, value=category.name)
        inc_cat_ws.cell(row=row, column=2, value=category.description)
        inc_cat_ws.cell(row=row, column=3, value=category.color)

    # Expenses worksheet
    exp_headers = [
        "Category",
        "Date",
        "Description",
        "Detailed Description",
        "Amount (cents)",
        "Created At",
    ]
    for col, header in enumerate(exp_headers, 1):
        cell = expenses_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, expense in enumerate(
        Expenses.objects.filter(user=user).order_by("-spent_at"), 2
    ):
        expenses_ws.cell(
            row=row, column=1, value=expense.category.name if expense.category else ""
        )
        expenses_ws.cell(row=row, column=2, value=expense.spent_at.strftime("%Y-%m-%d"))
        expenses_ws.cell(row=row, column=3, value=expense.description)
        expenses_ws.cell(row=row, column=4, value=expense.detailed_description)
        expenses_ws.cell(row=row, column=5, value=expense.amount)
        expenses_ws.cell(row=row, column=6, value=expense.created_at.isoformat())

    # Incomes worksheet
    inc_headers = [
        "Category",
        "Date",
        "Description",
        "Detailed Description",
        "Amount (cents)",
        "Created At",
    ]
    for col, header in enumerate(inc_headers, 1):
        cell = incomes_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, income in enumerate(
        Incomes.objects.filter(user=user).order_by("-received_at"), 2
    ):
        incomes_ws.cell(
            row=row, column=1, value=income.category.name if income.category else ""
        )
        incomes_ws.cell(
            row=row, column=2, value=income.received_at.strftime("%Y-%m-%d")
        )
        incomes_ws.cell(row=row, column=3, value=income.description)
        incomes_ws.cell(row=row, column=4, value=income.detailed_description)
        incomes_ws.cell(row=row, column=5, value=income.amount)
        incomes_ws.cell(row=row, column=6, value=income.created_at.isoformat())

    # Auto-adjust column widths
    for ws in [info_ws, exp_cat_ws, inc_cat_ws, expenses_ws, incomes_ws]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"fd_{user.username}_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
@require_POST
def import_financial_data(request):
    """Unified import function that handles JSON, CSV, and Excel files"""
    try:
        # Check if file was uploaded
        if "file" not in request.FILES:
            return JsonResponse(
                {"success": False, "message": "Nenhum arquivo foi enviado."}
            )

        uploaded_file = request.FILES["file"]
        filename = uploaded_file.name.lower()

        # Check file size (limit to 10MB)
        if uploaded_file.size > 10 * 1024 * 1024:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Arquivo muito grande. Limite máximo: 10MB.",
                }
            )

        # Route to appropriate handler based on file extension
        if filename.endswith(".json"):
            return _import_json_data(uploaded_file, request)
        elif filename.endswith(".csv"):
            return _import_csv_data(uploaded_file, request)
        elif filename.endswith((".xlsx", ".xls")):
            return _import_excel_data(uploaded_file, request)
        else:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Formato de arquivo não suportado. Use JSON, CSV ou Excel (.xlsx, .xls).",
                }
            )

    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro durante a importação: {str(e)}"}
        )


def _import_json_data(uploaded_file, request):
    """Handle JSON file import"""
    try:
        for i in range(20):
            print("_")
        file_content = uploaded_file.read().decode("utf-8")
        data = json.loads(file_content)
    except json.JSONDecodeError as e:
        return JsonResponse(
            {"success": False, "message": f"Arquivo JSON inválido. Erro: {str(e)}"}
        )
    except UnicodeDecodeError:
        return JsonResponse(
            {
                "success": False,
                "message": "Codificação do arquivo inválida. Use UTF-8.",
            }
        )

    # Validate that data is a dictionary
    if not isinstance(data, dict):
        return JsonResponse(
            {
                "success": False,
                "message": "Formato JSON inválido. O arquivo deve conter um objeto JSON.",
            }
        )

    # Create default structure if fields are missing
    required_fields = [
        "expense_categories",
        "income_categories",
        "expenses",
        "incomes",
    ]

    # Initialize missing fields with empty lists
    for field in required_fields:
        if field not in data:
            data[field] = []

    # Validate that required fields are lists
    for field in required_fields:
        if not isinstance(data[field], list):
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Campo '{field}' deve ser uma lista no arquivo JSON.",
                }
            )

    user = request.user
    clear_data = request.POST.get("clear_existing", "false").lower() == "true"

    return _process_import_data(data, user, clear_data)


def _import_csv_data(uploaded_file, request):
    """Handle CSV file import"""
    try:
        file_content = uploaded_file.read().decode("utf-8")
        csv_reader = csv.reader(file_content.splitlines())
        rows = list(csv_reader)
    except UnicodeDecodeError:
        return JsonResponse(
            {
                "success": False,
                "message": "Codificação do arquivo inválida. Use UTF-8.",
            }
        )

    if len(rows) < 5:
        return JsonResponse(
            {"success": False, "message": "Arquivo CSV muito pequeno ou inválido."}
        )

    user = request.user
    clear_data = request.POST.get("clear_existing", "false").lower() == "true"

    # Parse CSV structure
    data = {
        "expense_categories": [],
        "income_categories": [],
        "expenses": [],
        "incomes": [],
    }

    current_section = None
    headers = []

    for row in rows:
        if not row or not any(row):  # Skip empty rows
            continue

        # Check for section headers
        if len(row) >= 1:
            if row[0] == "EXPENSE CATEGORIES":
                current_section = "expense_categories"
                continue
            elif row[0] == "INCOME CATEGORIES":
                current_section = "income_categories"
                continue
            elif row[0] == "EXPENSES":
                current_section = "expenses"
                continue
            elif row[0] == "INCOMES":
                current_section = "incomes"
                continue

            # Check for headers row
            if current_section and row[0] in ["Name", "Category"]:
                headers = row
                continue

        # Process data rows
        if current_section and headers and len(row) >= len(headers):
            row_data = dict(zip(headers, row))

            if current_section == "expense_categories":
                data["expense_categories"].append(
                    {
                        "name": row_data.get("Name", ""),
                        "description": row_data.get("Description", ""),
                        "color": row_data.get("Color", "#FFFFFF"),
                    }
                )
            elif current_section == "income_categories":
                data["income_categories"].append(
                    {
                        "name": row_data.get("Name", ""),
                        "description": row_data.get("Description", ""),
                        "color": row_data.get("Color", "#FFFFFF"),
                    }
                )
            elif current_section == "expenses":
                data["expenses"].append(
                    {
                        "category": row_data.get("Category", ""),
                        "spent_at": row_data.get("Date", ""),
                        "description": row_data.get("Description", ""),
                        "detailed_description": row_data.get(
                            "Detailed Description", ""
                        ),
                        "amount": int(row_data.get("Amount (cents)", "0"))
                        if row_data.get("Amount (cents)", "").isdigit()
                        else 0,
                    }
                )
            elif current_section == "incomes":
                data["incomes"].append(
                    {
                        "category": row_data.get("Category", ""),
                        "received_at": row_data.get("Date", ""),
                        "description": row_data.get("Description", ""),
                        "detailed_description": row_data.get(
                            "Detailed Description", ""
                        ),
                        "amount": int(row_data.get("Amount (cents)", "0"))
                        if row_data.get("Amount (cents)", "").isdigit()
                        else 0,
                    }
                )

    return _process_import_data(data, user, clear_data)


def _import_excel_data(uploaded_file, request):
    """Handle Excel file import"""
    if not EXCEL_AVAILABLE:
        return JsonResponse(
            {
                "success": False,
                "message": "Importação Excel não disponível. Instale openpyxl: pip install openpyxl",
            }
        )

    try:
        wb = load_workbook(uploaded_file, read_only=True)
    except Exception:
        return JsonResponse(
            {"success": False, "message": "Arquivo Excel inválido ou corrompido."}
        )

    user = request.user
    clear_data = request.POST.get("clear_existing", "false").lower() == "true"

    # Parse Excel structure
    data = {
        "expense_categories": [],
        "income_categories": [],
        "expenses": [],
        "incomes": [],
    }

    # Process each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name == "Expense Categories":
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Check if first column has data
                    data["expense_categories"].append(
                        {
                            "name": str(row[0]) if row[0] else "",
                            "description": str(row[1]) if row[1] else "",
                            "color": str(row[2]) if row[2] else "#FFFFFF",
                        }
                    )

        elif sheet_name == "Income Categories":
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Check if first column has data
                    data["income_categories"].append(
                        {
                            "name": str(row[0]) if row[0] else "",
                            "description": str(row[1]) if row[1] else "",
                            "color": str(row[2]) if row[2] else "#FFFFFF",
                        }
                    )

        elif sheet_name == "Expenses":
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[1]:  # Check if date column has data
                    try:
                        # Handle date formatting
                        date_value = row[1]
                        if hasattr(date_value, "strftime"):
                            date_str = date_value.strftime("%Y-%m-%d")
                        else:
                            date_str = str(date_value)

                        data["expenses"].append(
                            {
                                "category": str(row[0]) if row[0] else "",
                                "spent_at": date_str,
                                "description": str(row[2]) if row[2] else "",
                                "detailed_description": str(row[3]) if row[3] else "",
                                "amount": int(row[4])
                                if row[4] and str(row[4]).replace(".", "").isdigit()
                                else 0,
                            }
                        )
                    except (ValueError, IndexError):
                        continue

        elif sheet_name == "Incomes":
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[1]:  # Check if date column has data
                    try:
                        # Handle date formatting
                        date_value = row[1]
                        if hasattr(date_value, "strftime"):
                            date_str = date_value.strftime("%Y-%m-%d")
                        else:
                            date_str = str(date_value)

                        data["incomes"].append(
                            {
                                "category": str(row[0]) if row[0] else "",
                                "received_at": date_str,
                                "description": str(row[2]) if row[2] else "",
                                "detailed_description": str(row[3]) if row[3] else "",
                                "amount": int(row[4])
                                if row[4] and str(row[4]).replace(".", "").isdigit()
                                else 0,
                            }
                        )
                    except (ValueError, IndexError):
                        continue

    return _process_import_data(data, user, clear_data)


def _process_import_data(data, user, clear_data):
    """Shared logic for processing import data from any format"""
    from django.db import transaction

    with transaction.atomic():
        if clear_data:
            # Clear existing data
            Expenses.objects.filter(user=user).delete()
            Incomes.objects.filter(user=user).delete()
            ExpenseCategory.objects.filter(user=user).delete()
            IncomeCategorys.objects.filter(user=user).delete()

        # Import expense categories
        expense_categories = {}
        categories_created = 0
        for cat_data in data.get("expense_categories", []):
            if not cat_data.get("name"):
                continue

            category, created = ExpenseCategory.objects.get_or_create(
                user=user,
                name=cat_data["name"],
                defaults={
                    "description": cat_data.get("description", ""),
                    "color": cat_data.get("color", "#FFFFFF"),
                },
            )
            expense_categories[cat_data["name"]] = category
            if created:
                categories_created += 1

        # Import income categories
        income_categories = {}
        income_categories_created = 0
        for cat_data in data.get("income_categories", []):
            if not cat_data.get("name"):
                continue

            category, created = IncomeCategorys.objects.get_or_create(
                user=user,
                name=cat_data["name"],
                defaults={
                    "description": cat_data.get("description", ""),
                    "color": cat_data.get("color", "#FFFFFF"),
                },
            )
            income_categories[cat_data["name"]] = category
            if created:
                income_categories_created += 1

        # Import expenses
        expenses_created = 0
        expenses_skipped = 0
        for expense_data in data.get("expenses", []):
            try:
                category = None
                if expense_data.get("category"):
                    category = expense_categories.get(expense_data["category"])

                spent_at = datetime.strptime(
                    expense_data["spent_at"], "%Y-%m-%d"
                ).date()

                Expenses.objects.create(
                    user=user,
                    category=category,
                    spent_at=spent_at,
                    description=expense_data.get("description", ""),
                    detailed_description=expense_data.get("detailed_description", ""),
                    amount=int(expense_data.get("amount", 0)),
                )
                expenses_created += 1
            except (ValueError, KeyError, TypeError):
                expenses_skipped += 1
                continue

        # Import incomes
        incomes_created = 0
        incomes_skipped = 0
        for income_data in data.get("incomes", []):
            try:
                category = None
                if income_data.get("category"):
                    category = income_categories.get(income_data["category"])

                received_at = datetime.strptime(
                    income_data["received_at"], "%Y-%m-%d"
                ).date()

                Incomes.objects.create(
                    user=user,
                    category=category,
                    received_at=received_at,
                    description=income_data.get("description", ""),
                    detailed_description=income_data.get("detailed_description", ""),
                    amount=int(income_data.get("amount", 0)),
                )
                incomes_created += 1
            except (ValueError, KeyError, TypeError):
                incomes_skipped += 1
                continue

    # Prepare success message
    total_created = (
        categories_created
        + income_categories_created
        + expenses_created
        + incomes_created
    )
    total_skipped = expenses_skipped + incomes_skipped

    message_parts = [
        "Importação concluída com sucesso!",
        f"✅ {categories_created} categorias de gastos",
        f"✅ {income_categories_created} categorias de receitas",
        f"✅ {expenses_created} gastos",
        f"✅ {incomes_created} receitas",
    ]

    if total_skipped > 0:
        message_parts.append(
            f"⚠️ {total_skipped} registros ignorados por dados inválidos"
        )

    return JsonResponse(
        {
            "success": True,
            "message": "\n".join(message_parts),
            "stats": {
                "expense_categories": categories_created,
                "income_categories": income_categories_created,
                "expenses": expenses_created,
                "incomes": incomes_created,
                "total_created": total_created,
                "total_skipped": total_skipped,
            },
        }
    )
